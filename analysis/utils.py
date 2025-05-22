import os
import pandas as pd
import shutil
from openpyxl import load_workbook

def convert_time(seconds
                 , time_unit):
    """Convert seconds into the specified time unit."""
    """Parameters:
        seconds (int): time duration in seconds.
        time_unit (str): desired output unit ('sec', 'min', 'hr', 'day', 'week', 'month')."""
    """Returns: (int): converted time in the specified unit, or -1 if unit is invalid."""
    
    time_unit = time_unit.lower()
    if time_unit == 'sec':
        return seconds
    if time_unit == 'min':
        return seconds//60
    if time_unit == 'hr':
        return seconds//(60*60)
    if time_unit == 'day':
        return seconds//(60*60*24)
    if time_unit == 'week':
        return seconds//(60*60*24*7)
    if time_unit == 'month':
        return seconds//(60*60*24*30)
    
    return -1


def add_timestamp_diffs(df
                   , timestamp_col = 'timestamp_int'
                   , time_unit = 'min'):
    """Add time difference columns to a DataFrame."""
    """Parameters:
        df (pd.DataFrame): dataframe with timestamp column.
        timestamp_col (str): name of the column with Unix timestamps.
        time_unit (str): time unit to convert differences to."""
    """Returns: (pd.DataFrame): dataframe with additional columns:
            - 'timestamp_diff': difference in seconds
            - 'time_diff_<time_unit>': converted time difference"""
    df['timestamp_diff'] = df[timestamp_col].diff().fillna(0).astype(int)
    df['time_diff_' + time_unit] = convert_time(df['timestamp_diff'], time_unit)
    return df


def add_reset_trade_count(df):
    """Add a column counting trades between resets."""
    """Parameters:
        df (pd.DataFrame): dataframe with 'recent_reset' and 'last_reset' columns."""
    """Returns: (pd.DataFrame): dataframe with a new 'trade_count' column."""
    df['trade_count'] = df['recent_reset'] - df['last_reset']

    
def filter_by_timestamp_range(primary_df, secondary_df, timestamp_col_pr='date_time', timestamp_col_sec='timestamp_floored'):
    """Filters secondary_df to only include rows with timestamps that fall within 
        the timestamp range defined by primary_df."""
    """Parameters:
        primary_df (pd.DataFrame): dataframe to be filtered.
        secondary_df (pd.DataFrame): dataframe providing the time window.
        timestamp_col_pr (str): name of the timestamp column in the primary dataframe.
        timestamp_col_sec (str): name of the timestamp column in the secondary dataframe."""
    """Returns: (pd.DataFrame): filtered version of secondary_df."""
    
    # Get time window from primary_df
    min_time = primary_df[timestamp_col_pr].min()
    max_time = primary_df[timestamp_col_pr].max()
    
    # Filter secondary_df within time window
    filtered_df = secondary_df[(secondary_df[timestamp_col_sec] >= min_time) & (secondary_df[timestamp_col_sec] <= max_time)].copy()
    
    return filtered_df.reset_index(drop=True)

def get_data_path(file_name
                  , data_type):
    """Construct and return the full path to a data file based on its type."""
    """Parameters:
        file_name (str): name of the data file (e.g., 'model_predictions.xlsx').
        data_type (str): type of data ('model' or 'trade')."""
    """Returns: (str): full path to the requested file, or None if data_type is invalid."""
    
    # Build absolute path to the project's 'data' directory 
    data_folder_path = os.path.join(os.path.abspath(os.path.join(os.getcwd(), '..')), 'data')
    # Return full path to file inside the appropriate subfolder
    if data_type == 'model':
        return os.path.join(data_folder_path, 'model_data', file_name)
    if data_type == 'trade':
        return os.path.join(data_folder_path, 'trading_data', file_name)
    # Return None if the data_type is not recognized
    return None


def get_closest_sheet_name(minutes):
    """Given a number of minutes, return the closest matching sheet name
    from the set ['15M', '30M', '60M', '90M']."""
    """Parameters:
        minutes (int or float): Desired time interval in minutes."""
    """Returns:(str): Closest matching sheet name (e.g., '15M')."""
    
    # If new options are added to the model spreadsheet, update this list
    sheet_options = [15, 30, 60, 90]
    # Find the closest time interval
    closest = min(sheet_options, key=lambda x: abs(x - minutes))
    return f"{closest}M"

def remove_extension(file_name):
    """Removes the extension part from the filename."""
    """Parameters:
        file_name (str): file name that needs extension removed"""
    """Returns:(str): cleaned file name if extension is defined or unedited file name"""
    
    extensions = ['.csv', '.xlsx']
    for ext in extensions:
        if file_name[len(file_name)-len(ext):] == '.csv':
            return file_name[:len(file_name)-len(ext)]
    
    return file_name

def fill_template_with_data(df
                            , file_name
                            , template_name='template_dsf_model.xlsx'):
    """
    Create a copy of a template Excel file and fill the 'Data' sheet starting from row 2 
    with selected columns from the input DataFrame.

    Parameters:
        df (pd.DataFrame): DataFrame containing columns:
            ['block_number', 'tx_hash', 'token_amount_a', 'token_amount_b', 'timestamp']
        filename (str): Name of the new Excel file to create.
        template_path (str): Path to the template Excel file (default: 'template_dsf_model.xlsx').

    Returns:
        None
    """
    # Ensure required columns are present
    required_cols = ['block_number', 'tx_hash', 'token_amount_a', 'token_amount_b', 'timestamp_int']
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(f"DataFrame is missing required columns: {missing}")
    
    template_path = get_data_path(template_name, 'model')
    file_path = get_data_path(file_name, 'model')
    # Copy the template to new file
    shutil.copy(template_path, file_path)

    # Load workbook and sheet
    wb = load_workbook(file_path)
    ws = wb['Data']
    cols_to_write = [1,3,4,5,6]
    # Write rows starting from row 2 (index 0-based = row 1)
    for i in range(len(df)):
        for j in range(len(required_cols)):
            ws.cell(row=i+2, column=cols_to_write[j], value=df.loc[i, required_cols[j]])

    # Save the modified file
    wb.save(file_path)